import os
import openpyxl
from tkinter import Tk, filedialog, Label
import re

def get_details_from_file(file_path):
    with open(file_path, 'r') as file:
        data = file.read()
    details = re.findall(r'name\n(.*?)\npassport number\n(.*?)\nticket number\n(.*?)\ntrade\n(.*?)\ndate\n(.*?)\n', data, re.DOTALL)
    if details:
        name, passport_number, ticket_number, trade, date = details[0]
        if all(detail.strip() for detail in details[0]):  # Check if all details are present and not empty
            return name.strip(), passport_number.strip(), ticket_number.strip(), trade.strip(), date.strip()
    return None

def traverse_and_collect_details(parent_folder):
    details_list = []
    incomplete_files = 0
    for root, dirs, files in os.walk(parent_folder):
        for file in files:
            if file == 'details.txt':
                file_path = os.path.join(root, file)
                details = get_details_from_file(file_path)
                if details:
                    details_list.append(details)
                else:
                    incomplete_files += 1
    return details_list, incomplete_files

def append_to_excel(file_path, start_row, details_list):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    current_row = start_row
    for details in details_list:
        ws.cell(row=current_row, column=2).value = details[0]
        ws.cell(row=current_row, column=3).value = details[1]
        ws.cell(row=current_row, column=4).value = details[2]
        ws.cell(row=current_row, column=5).value = details[3]
        date_value = details[4]
        if re.match(r'\d{2}\.\d{2}\.\d{4}', date_value):  # Check if date is in dd.mm.yyyy format
            day, month, year = date_value.split('.')
            ws.cell(row=current_row, column=6).value = f'=DATE({year},{month},{day})'
        else:
            ws.cell(row=current_row, column=6).value = date_value
        current_row += 1

    wb.save(file_path)

def show_incomplete_files_and_appended_count(incomplete_files_count, appended_entries_count):
    # Initialize Tkinter window
    window = Tk()
    window.title("Incomplete Files and Appended Entries Count")
    
    incomplete_label = Label(window, text=f"Number of incomplete files: {incomplete_files_count}", font=("Helvetica", 14))
    incomplete_label.pack(padx=20, pady=10)

    appended_label = Label(window, text=f"Number of entries appended to Excel: {appended_entries_count}", font=("Helvetica", 14))
    appended_label.pack(padx=20, pady=10)
    
    window.mainloop()

def main():
    # Initialize Tkinter
    root = Tk()
    root.withdraw()  # Hide the main Tkinter window

    current_folder = os.getcwd()
    excel_file = filedialog.askopenfilename(initialdir=current_folder, title="Select Excel File", filetypes=[("Excel files", "*.xlsx")])
    if not excel_file:
        print("No Excel file selected. Exiting.")
        return

    start_row = int(input("Enter the start row number: "))

    details_list, incomplete_files = traverse_and_collect_details(current_folder)
    appended_entries_count = len(details_list)

    if details_list:
        append_to_excel(excel_file, start_row, details_list)
        print("Details appended to the Excel file successfully.")
    else:
        print("No valid details.txt files found with all required details.")

    show_incomplete_files_and_appended_count(incomplete_files, appended_entries_count)

if __name__ == "__main__":
    main()
